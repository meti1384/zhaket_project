import openpyxl

def transfer_data_form_excel(path):
    file_path = path

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    data = []
    headers = worksheet[1]
    header_names = [cell.value for cell in headers]

    num_rows=(worksheet.max_row)  
    workbook.close() 

    number_of_rows=[f"row{i}" for i in range (1,num_rows)]
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for header, cell_value in zip(header_names, row):
            row_data[header] = cell_value 
        data.append(row_data)
    return(dict(zip(number_of_rows,data)))

path_file=r"C:\Users\mahdi\Downloads\Telegram Desktop\updater (2).xlsx"
main_data=transfer_data_form_excel(path_file)
print(main_data)


